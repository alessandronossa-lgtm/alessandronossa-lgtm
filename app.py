import os
import re
import json
import tempfile
import unicodedata
from decimal import Decimal, InvalidOperation
from datetime import datetime, timedelta, timezone
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, jsonify,
    session, url_for, send_file, abort
)

import mercadopago
import requests
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# =====================================================
# APP CONFIG
# =====================================================

app = Flask(__name__)

SECRET_KEY = os.getenv("SECRET_KEY")
if not SECRET_KEY:
    raise Exception("SECRET_KEY não configurada.")
app.secret_key = SECRET_KEY

DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    raise Exception("DATABASE_URL não configurada.")

if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {"pool_pre_ping": True}

db = SQLAlchemy(app)

MP_ACCESS_TOKEN = os.getenv("MP_ACCESS_TOKEN")
if not MP_ACCESS_TOKEN:
    raise Exception("MP_ACCESS_TOKEN não configurado.")
sdk = mercadopago.SDK(MP_ACCESS_TOKEN)

ADMIN_EMAIL = (os.getenv("ADMIN_EMAIL") or "promptsheetbrasil@gmail.com").strip().lower()


def now_utc():
    return datetime.now(timezone.utc)


def base_url():
    env = os.getenv("BASE_URL")
    if env and env.strip():
        return env.rstrip("/")
    return request.host_url.rstrip("/")


# =====================================================
# MODELS
# =====================================================

class Config(db.Model):
    __tablename__ = "config"

    id = db.Column(db.Integer, primary_key=True)
    price_avulso_24h = db.Column(db.Numeric(10, 2), nullable=False, default=Decimal("9.90"))
    price_premium_mensal = db.Column(db.Numeric(10, 2), nullable=False, default=Decimal("19.90"))
    updated_at = db.Column(db.DateTime(timezone=True), default=now_utc, onupdate=now_utc)


class Usuario(db.Model):
    __tablename__ = "usuario"

    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(150), unique=True, nullable=False)
    senha_hash = db.Column(db.String(255), nullable=False)

    subscription_status = db.Column(db.String(50), default="none")
    subscription_id = db.Column(db.String(120), nullable=True)
    free_premium_until = db.Column(db.DateTime(timezone=True), nullable=True)

    created_at = db.Column(db.DateTime(timezone=True), default=now_utc)

    def set_senha(self, senha: str):
        self.senha_hash = generate_password_hash(senha)

    def verificar_senha(self, senha: str) -> bool:
        return check_password_hash(self.senha_hash, senha)

    def premium_ativo(self) -> bool:
        status = (self.subscription_status or "").lower()
        if status in ("authorized", "active"):
            return True
        if self.free_premium_until and self.free_premium_until > now_utc():
            return True
        return False

    def is_admin(self) -> bool:
        return (self.email or "").strip().lower() == ADMIN_EMAIL


class Projeto(db.Model):
    __tablename__ = "projeto"

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("usuario.id"), nullable=False)
    nome = db.Column(db.String(120), nullable=False)
    prompt = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime(timezone=True), default=now_utc)


class AcessoProjeto(db.Model):
    __tablename__ = "acesso_projeto"

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("usuario.id"), nullable=False)
    projeto_id = db.Column(db.Integer, db.ForeignKey("projeto.id"), nullable=False)
    expires_at = db.Column(db.DateTime(timezone=True), nullable=False)
    payment_id = db.Column(db.String(120), nullable=True)
    created_at = db.Column(db.DateTime(timezone=True), default=now_utc)


with app.app_context():
    db.create_all()


# =====================================================
# HELPERS GERAIS
# =====================================================

def get_config() -> Config:
    cfg = db.session.get(Config, 1)
    if not cfg:
        cfg = Config(id=1, price_avulso_24h=Decimal("9.90"), price_premium_mensal=Decimal("19.90"))
        db.session.add(cfg)
        db.session.commit()
    return cfg


def get_prices():
    cfg = get_config()
    return float(cfg.price_avulso_24h), float(cfg.price_premium_mensal)


def current_user():
    uid = session.get("user_id")
    if not uid:
        return None
    return db.session.get(Usuario, uid)


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        u = current_user()
        if not u:
            return redirect(url_for("login"))
        if not u.is_admin():
            abort(403)
        return fn(*args, **kwargs)
    return wrapper


def sanitize_project_filename(name: str) -> str:
    safe = secure_filename((name or "planilha").strip())
    return safe or "planilha"


def to_decimal(value, fallback: Decimal) -> Decimal:
    try:
        return Decimal(str(value).replace(",", ".").strip())
    except (InvalidOperation, AttributeError, ValueError):
        return fallback


# =====================================================
# MOTOR DE GERAÇÃO DE PLANILHA
# =====================================================

def normalize_text(text: str) -> str:
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return text.lower().strip()


def clean_column_name(name: str) -> str:
    name = re.sub(r"\s+", " ", (name or "").strip(" -,:;|"))
    if not name:
        return ""
    return name[:40]


def unique_columns(columns):
    seen = set()
    final = []
    for col in columns:
        key = normalize_text(col)
        if key and key not in seen:
            seen.add(key)
            final.append(col)
    return final


def split_candidate_columns(text: str):
    text = (text or "").replace("\n", ", ")
    text = re.sub(r"\s+e\s+", ", ", text, flags=re.IGNORECASE)
    text = text.replace(";", ",").replace("|", ",")
    parts = [clean_column_name(p) for p in text.split(",")]
    return [p for p in parts if p]


def extract_explicit_columns(prompt: str):
    if not prompt:
        return []

    prompt_clean = prompt.strip()

    patterns = [
        r"colunas?\s*[:\-]\s*(.+)",
        r"deve ter\s+colunas?\s*[:\-]?\s*(.+)",
        r"nela deve ter\s+colunas?\s*[:\-]?\s*(.+)",
        r"campos?\s*[:\-]\s*(.+)",
    ]

    for pattern in patterns:
        match = re.search(pattern, prompt_clean, flags=re.IGNORECASE)
        if match:
            tail = match.group(1)
            tail = re.split(r"\.|\n", tail)[0]
            cols = split_candidate_columns(tail)
            if cols:
                return unique_columns([c.title() for c in cols])

    found = re.findall(r"coluna\s+([a-zA-ZÀ-ÿ0-9 _/-]+)", prompt_clean, flags=re.IGNORECASE)
    if found:
        cols = []
        for item in found:
            item = re.split(r",|\.|\n", item)[0]
            item = clean_column_name(item)
            if item:
                cols.append(item.title())
        return unique_columns(cols)

    return []


def infer_columns_from_prompt(prompt: str):
    p = normalize_text(prompt)

    if any(k in p for k in ["estoque", "inventario", "almoxarifado"]):
        return ["Data", "Produto", "Categoria", "Quantidade", "Estoque Mínimo", "Fornecedor"]

    if any(k in p for k in ["venda", "vendas", "comissao", "comissão", "faturamento"]):
        return ["Data", "Cliente", "Produto", "Quantidade", "Preço Unitário", "Valor Total"]

    if any(k in p for k in ["financeiro", "fluxo de caixa", "caixa"]):
        return ["Data", "Descrição", "Categoria", "Entrada", "Saída", "Saldo"]

    if any(k in p for k in ["despesa", "despesas", "gastos", "custos"]):
        return ["Data", "Descrição", "Categoria", "Valor", "Forma de Pagamento", "Observação"]

    if any(k in p for k in ["pedido", "pedidos", "orcamento", "orçamento"]):
        return ["Data", "Cliente", "Produto", "Quantidade", "Preço Unitário", "Valor Total", "Status"]

    if any(k in p for k in ["cliente", "clientes", "cadastro de clientes"]):
        return ["Nome", "Telefone", "E-mail", "Cidade", "Observação"]

    if any(k in p for k in ["funcionario", "funcionário", "colaborador", "equipe"]):
        return ["Nome", "Cargo", "Telefone", "E-mail", "Cidade", "Observação"]

    return ["Data", "Descrição", "Valor"]


def detect_columns(prompt: str):
    explicit = extract_explicit_columns(prompt)
    if explicit:
        return explicit
    return infer_columns_from_prompt(prompt)


def is_money_column(name: str) -> bool:
    n = normalize_text(name)
    keys = ["preco", "valor", "entrada", "saida", "saldo", "custo", "total"]
    return any(k in n for k in keys)


def is_integer_column(name: str) -> bool:
    n = normalize_text(name)
    keys = ["quantidade", "qtd", "estoque", "minimo"]
    return any(k in n for k in keys)


def has_column(columns, target):
    target_n = normalize_text(target)
    return any(normalize_text(c) == target_n for c in columns)


def col_index(columns, target):
    target_n = normalize_text(target)
    for idx, col in enumerate(columns, start=1):
        if normalize_text(col) == target_n:
            return idx
    return None


def value_for_column(col_name: str, row_number: int):
    name = normalize_text(col_name)

    if name == "data":
        return f"0{row_number}/04/2026"
    if name == "cliente":
        return f"Cliente Exemplo {row_number - 5}"
    if name == "produto":
        return f"Produto {chr(64 + row_number - 5)}"
    if name == "categoria":
        return f"Categoria {row_number - 5}"
    if name == "fornecedor":
        return "Fornecedor Exemplo"
    if name == "status":
        return "Pendente" if row_number == 6 else "Pago"
    if name == "descricao":
        return "Lançamento exemplo" if row_number == 6 else "Movimentação exemplo"
    if name == "observacao":
        return ""
    if name == "forma de pagamento":
        return "PIX" if row_number == 6 else "Boleto"
    if name == "telefone":
        return "(27) 99999-9999"
    if name == "e-mail":
        return f"contato{row_number - 5}@exemplo.com"
    if name == "cidade":
        return "Vila Velha"
    if name == "nome":
        return f"Nome Exemplo {row_number - 5}"
    if name == "cargo":
        return "Vendedor"

    if name in ("quantidade", "qtd"):
        return 2 if row_number == 6 else 3
    if name == "estoque minimo":
        return 10 if row_number == 6 else 5
    if name == "estoque":
        return 25 if row_number == 6 else 8
    if name == "preco unitario":
        return 35.0 if row_number == 6 else 20.0
    if name == "valor":
        return 150.0 if row_number == 6 else 90.0
    if name == "entrada":
        return 500.0 if row_number == 6 else 0.0
    if name == "saida":
        return 0.0 if row_number == 6 else 120.0

    if name in ("valor total", "saldo"):
        return None

    return ""


def style_sheet(ws, columns, prompt, project_name):
    blue = "1F4E78"
    dark = "0F243E"
    light_fill = "F7FBFF"
    total_fill = "FFF2CC"
    white = "FFFFFF"

    thin_gray = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    title_end_col = max(1, min(4, len(columns)))
    subtitle_end_col = max(1, min(6, len(columns)))
    prompt_end_col = max(1, min(max(6, len(columns)), 8))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=title_end_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=subtitle_end_col)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=prompt_end_col)

    ws["A1"] = "PromptSheet"
    ws["A1"].font = Font(size=18, bold=True, color=dark)
    ws["A2"] = f"Projeto: {project_name}"
    ws["A2"].font = Font(size=11, bold=True, color=blue)
    ws["A3"] = f"Prompt: {prompt or ''}"
    ws["A3"].font = Font(size=10, italic=True, color="666666")
    ws["A3"].alignment = Alignment(wrap_text=True)

    header_row = 5
    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=idx, value=col)
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:{get_column_letter(len(columns))}5"

    for idx, col in enumerate(columns, start=1):
        width = max(len(col) + 4, 14)
        ws.column_dimensions[get_column_letter(idx)].width = min(width, 28)

    for row in range(6, 50):
        for col in range(1, len(columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=light_fill)

    for idx, col in enumerate(columns, start=1):
        if is_money_column(col):
            for row in range(6, 200):
                ws.cell(row=row, column=idx).number_format = 'R$ #,##0.00'
        elif is_integer_column(col):
            for row in range(6, 200):
                ws.cell(row=row, column=idx).number_format = '0'

    total_row = 8
    ws.cell(row=total_row, column=1, value="Totais")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor=total_fill)
    ws.cell(row=total_row, column=1).border = border

    for idx, col in enumerate(columns, start=2):
        n = normalize_text(col)
        letter = get_column_letter(idx)
        cell = ws.cell(row=total_row, column=idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=total_fill)
        cell.border = border

        if any(k in n for k in ["quantidade", "valor", "entrada", "saida", "preco", "custo", "estoque"]):
            cell.value = f"=SUM({letter}6:{letter}7)"
        elif n == "saldo":
            cell.value = f"={letter}7"

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[3].height = 34
    ws.row_dimensions[5].height = 24

    return ws


def apply_formulas(ws, columns):
    q_idx = col_index(columns, "Quantidade")
    pu_idx = col_index(columns, "Preço Unitário")
    vt_idx = col_index(columns, "Valor Total")

    if q_idx and pu_idx and vt_idx:
        for row in (6, 7):
            q_letter = get_column_letter(q_idx)
            pu_letter = get_column_letter(pu_idx)
            ws.cell(row=row, column=vt_idx).value = f"={q_letter}{row}*{pu_letter}{row}"

    e_idx = col_index(columns, "Entrada")
    s_idx = col_index(columns, "Saída")
    saldo_idx = col_index(columns, "Saldo")

    if e_idx and s_idx and saldo_idx:
        e_letter = get_column_letter(e_idx)
        s_letter = get_column_letter(s_idx)
        saldo_letter = get_column_letter(saldo_idx)

        ws.cell(row=6, column=saldo_idx).value = f"={e_letter}6-{s_letter}6"
        ws.cell(row=7, column=saldo_idx).value = f"={saldo_letter}6+{e_letter}7-{s_letter}7"


def fill_example_data(ws, columns):
    for row in (6, 7):
        for idx, col_name in enumerate(columns, start=1):
            value = value_for_column(col_name, row)
            if value is not None:
                ws.cell(row=row, column=idx, value=value)

    apply_formulas(ws, columns)

    for idx, col in enumerate(columns, start=1):
        letter = get_column_letter(idx)
        max_len = len(str(col))
        for row in range(6, 9):
            val = ws.cell(row=row, column=idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[letter].width = min(max(max_len + 3, 14), 30)


def generate_workbook_from_prompt(project_name: str, prompt: str) -> Workbook:
    columns = detect_columns(prompt)

    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha"

    style_sheet(ws, columns, prompt, project_name)
    fill_example_data(ws, columns)

    return wb


# =====================================================
# HEALTH
# =====================================================

@app.route("/healthz")
def healthz():
    return "ok", 200


# =====================================================
# AUTH ROUTES
# =====================================================

@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        senha = request.form.get("senha") or ""

        if not email or not senha:
            return render_template("register.html", erro="Preencha email e senha.")

        if Usuario.query.filter_by(email=email).first():
            return render_template("register.html", erro="Esse email já está cadastrado.")

        u = Usuario(email=email)
        u.set_senha(senha)
        db.session.add(u)
        db.session.commit()

        session["user_id"] = u.id
        return redirect(url_for("app_home"))

    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        senha = request.form.get("senha") or ""

        u = Usuario.query.filter_by(email=email).first()
        if not u or not u.verificar_senha(senha):
            return render_template("login.html", erro="Email ou senha inválidos.")

        session["user_id"] = u.id
        return redirect(url_for("app_home"))

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("index"))


# =====================================================
# PÁGINAS
# =====================================================

@app.route("/")
def index():
    u = current_user()
    if u:
        return redirect(url_for("app_home"))
    return render_template("index.html", usuario=u)


@app.route("/app", methods=["GET", "POST"])
@login_required
def app_home():
    u = current_user()

    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        prompt = (request.form.get("prompt") or "").strip()

        if not nome:
            projetos = Projeto.query.filter_by(user_id=u.id).order_by(Projeto.created_at.desc()).all()
            return render_template("app.html", usuario=u, projetos=projetos, erro="Dê um nome ao projeto.")

        p = Projeto(user_id=u.id, nome=nome, prompt=prompt)
        db.session.add(p)
        db.session.commit()
        return redirect(url_for("projeto_view", projeto_id=p.id))

    projetos = Projeto.query.filter_by(user_id=u.id).order_by(Projeto.created_at.desc()).all()
    return render_template("app.html", usuario=u, projetos=projetos)


@app.route("/projeto/<int:projeto_id>")
@login_required
def projeto_view(projeto_id):
    u = current_user()
    p = Projeto.query.filter_by(id=projeto_id, user_id=u.id).first_or_404()

    premium = u.premium_ativo()
    avulso_price, premium_price = get_prices()

    acesso = AcessoProjeto.query.filter_by(
        user_id=u.id, projeto_id=p.id
    ).order_by(AcessoProjeto.expires_at.desc()).first()

    acesso_ativo = False
    expira_em = None
    if acesso and acesso.expires_at > now_utc():
        acesso_ativo = True
        expira_em = acesso.expires_at

    return render_template(
        "projeto.html",
        usuario=u,
        projeto=p,
        premium=premium,
        acesso_ativo=acesso_ativo,
        expira_em=expira_em,
        price_avulso=avulso_price,
        price_premium=premium_price
    )


# =====================================================
# PREMIUM
# =====================================================

@app.route("/premium")
@login_required
def premium_page():
    u = current_user()
    _, premium_price = get_prices()
    return render_template("premium.html", usuario=u, price_premium=premium_price)


@app.route("/premium/assinar", methods=["POST"])
@login_required
def premium_assinar():
    u = current_user()

    if u.premium_ativo():
        return jsonify({"erro": "Você já possui Premium ativo."}), 400

    _, premium_price = get_prices()

    url = "https://api.mercadopago.com/preapproval"
    headers = {
        "Authorization": f"Bearer {MP_ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }

    payload = {
        "reason": "PromptSheet Premium - Acesso ilimitado",
        "payer_email": u.email,
        "auto_recurring": {
            "frequency": 1,
            "frequency_type": "months",
            "transaction_amount": float(premium_price),
            "currency_id": "BRL"
        },
        "back_url": f"{base_url()}/premium/retorno",
        "external_reference": f"user:{u.id}"
    }

    r = requests.post(url, headers=headers, data=json.dumps(payload), timeout=30)
    if r.status_code >= 400:
        return jsonify({"erro": "Falha ao criar assinatura.", "detalhes": r.text}), 500

    resp = r.json()
    init_point = resp.get("init_point")
    sub_id = resp.get("id")

    if not init_point or not sub_id:
        return jsonify({"erro": "Resposta inesperada do Mercado Pago.", "detalhes": resp}), 500

    u.subscription_id = sub_id
    u.subscription_status = (resp.get("status") or "pending").lower()
    db.session.commit()

    return jsonify({"init_point": init_point})


@app.route("/premium/retorno")
@login_required
def premium_retorno():
    return redirect(url_for("app_home"))


# =====================================================
# AVULSO 24H
# =====================================================

@app.route("/projeto/<int:projeto_id>/comprar_diaria", methods=["POST"])
@login_required
def comprar_diaria(projeto_id):
    u = current_user()
    p = Projeto.query.filter_by(id=projeto_id, user_id=u.id).first_or_404()

    if u.premium_ativo():
        return jsonify({"erro": "Você já é Premium."}), 400

    avulso_price, _ = get_prices()
    external_reference = f"user:{u.id}|project:{p.id}|kind:daily"

    preference_data = {
        "items": [{
            "title": f"Acesso 24h - Projeto {p.nome}",
            "quantity": 1,
            "currency_id": "BRL",
            "unit_price": float(avulso_price)
        }],
        "payer": {"email": u.email},
        "external_reference": external_reference,
        "notification_url": f"{base_url()}/webhook",
        "back_urls": {
            "success": f"{base_url()}/pendente/{p.id}",
            "failure": f"{base_url()}/pendente/{p.id}",
            "pending": f"{base_url()}/pendente/{p.id}",
        },
        "auto_return": "all"
    }

    response = sdk.preference().create(preference_data)
    init_point = response.get("response", {}).get("init_point")

    if not init_point:
        return jsonify({"erro": "Não foi possível gerar o link de pagamento.", "detalhes": response}), 500

    return jsonify({"init_point": init_point})


@app.route("/pendente/<int:projeto_id>")
@login_required
def pendente_page(projeto_id):
    u = current_user()
    p = Projeto.query.filter_by(id=projeto_id, user_id=u.id).first_or_404()
    return render_template("pendente.html", projeto=p)


@app.route("/status")
@login_required
def status():
    u = current_user()
    projeto_id = request.args.get("projeto_id", type=int)

    if u.premium_ativo():
        return jsonify({"paid": True, "premium": True})

    if not projeto_id:
        return jsonify({"paid": False, "premium": False})

    acesso = AcessoProjeto.query.filter_by(
        user_id=u.id, projeto_id=projeto_id
    ).order_by(AcessoProjeto.expires_at.desc()).first()

    if acesso and acesso.expires_at > now_utc():
        return jsonify({"paid": True, "premium": False, "expires_at": acesso.expires_at.isoformat()})

    return jsonify({"paid": False, "premium": False})


# =====================================================
# DOWNLOAD
# =====================================================

@app.route("/projeto/<int:projeto_id>/download")
@login_required
def download_projeto(projeto_id):
    u = current_user()
    p = Projeto.query.filter_by(id=projeto_id, user_id=u.id).first_or_404()

    if not u.premium_ativo():
        acesso = AcessoProjeto.query.filter_by(
            user_id=u.id, projeto_id=p.id
        ).order_by(AcessoProjeto.expires_at.desc()).first()

        if not acesso or acesso.expires_at <= now_utc():
            return "Acesso negado. Compre a diária (24h) ou assine o Premium.", 403

    wb = generate_workbook_from_prompt(p.nome, p.prompt or "")
    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp.name)

    download_name = f"{sanitize_project_filename(p.nome)}.xlsx"
    return send_file(temp.name, as_attachment=True, download_name=download_name)


# =====================================================
# WEBHOOK
# =====================================================

def parse_webhook_event():
    data = request.get_json(silent=True) or {}

    if isinstance(data, dict) and data.get("type") and isinstance(data.get("data"), dict) and data["data"].get("id"):
        return data.get("type"), str(data["data"]["id"])

    t = request.args.get("type")
    did = request.args.get("data.id")
    if t and did:
        return t, str(did)

    topic = request.args.get("topic")
    _id = request.args.get("id")
    if topic and _id:
        return topic, str(_id)

    return None, None


def handle_payment(payment_id: str):
    pay_resp = sdk.payment().get(payment_id)
    payment = pay_resp.get("response", {})
    status = (payment.get("status") or "").lower()

    if status != "approved":
        return

    external_reference = payment.get("external_reference") or ""

    if "kind:daily" in external_reference and "project:" in external_reference and "user:" in external_reference:
        try:
            parts = external_reference.split("|")
            uid = int(parts[0].split(":")[1])
            pid = int(parts[1].split(":")[1])
        except Exception:
            return

        user = db.session.get(Usuario, uid)
        proj = db.session.get(Projeto, pid)
        if not user or not proj or proj.user_id != user.id:
            return

        existing_same_payment = AcessoProjeto.query.filter_by(
            user_id=user.id,
            projeto_id=proj.id,
            payment_id=str(payment_id)
        ).first()
        if existing_same_payment:
            return

        expires = now_utc() + timedelta(hours=24)

        acesso = AcessoProjeto.query.filter_by(
            user_id=user.id, projeto_id=proj.id
        ).order_by(AcessoProjeto.expires_at.desc()).first()

        if acesso and acesso.expires_at > now_utc():
            expires = acesso.expires_at + timedelta(hours=24)

        novo = AcessoProjeto(
            user_id=user.id,
            projeto_id=proj.id,
            expires_at=expires,
            payment_id=str(payment_id)
        )
        db.session.add(novo)
        db.session.commit()


def handle_preapproval(preapproval_id: str):
    url = f"https://api.mercadopago.com/preapproval/{preapproval_id}"
    headers = {"Authorization": f"Bearer {MP_ACCESS_TOKEN}"}
    r = requests.get(url, headers=headers, timeout=30)
    if r.status_code >= 400:
        return

    sub = r.json()
    status = (sub.get("status") or "").lower()
    ext = sub.get("external_reference") or ""

    uid = None
    if ext.startswith("user:"):
        try:
            uid = int(ext.split(":")[1])
        except Exception:
            uid = None

    user = db.session.get(Usuario, uid) if uid else None
    if not user:
        payer_email = (sub.get("payer_email") or "").strip().lower()
        if payer_email:
            user = Usuario.query.filter_by(email=payer_email).first()

    if not user:
        return

    user.subscription_id = preapproval_id
    user.subscription_status = status or "unknown"
    db.session.commit()


@app.route("/webhook", methods=["POST", "GET"])
def webhook():
    try:
        event_type, event_id = parse_webhook_event()

        if not event_type or not event_id:
            return jsonify({"status": "ignored"}), 200

        if event_type == "payment":
            handle_payment(event_id)
            return jsonify({"status": "ok"}), 200

        if event_type in ("preapproval",):
            handle_preapproval(event_id)
            return jsonify({"status": "ok"}), 200

        return jsonify({"status": "ignored"}), 200

    except Exception as e:
        print("Erro webhook:", e)
        db.session.rollback()
        return jsonify({"erro": "erro interno"}), 500


# =====================================================
# ADMIN
# =====================================================

@app.route("/admin")
@login_required
@admin_required
def admin_home():
    cfg = get_config()
    users_count = Usuario.query.count()
    projects_count = Projeto.query.count()
    return render_template(
        "admin.html",
        config=cfg,
        users_count=users_count,
        projects_count=projects_count,
        admin_email=ADMIN_EMAIL
    )


@app.route("/admin/users")
@login_required
@admin_required
def admin_users():
    users = Usuario.query.order_by(Usuario.created_at.desc()).all()
    return render_template("admin_users.html", users=users, admin_email=ADMIN_EMAIL)


@app.route("/admin/config", methods=["POST"])
@login_required
@admin_required
def admin_save_config():
    cfg = get_config()

    av = to_decimal(request.form.get("price_avulso_24h"), Decimal(str(cfg.price_avulso_24h)))
    pr = to_decimal(request.form.get("price_premium_mensal"), Decimal(str(cfg.price_premium_mensal)))

    cfg.price_avulso_24h = av.quantize(Decimal("0.01"))
    cfg.price_premium_mensal = pr.quantize(Decimal("0.01"))
    db.session.commit()

    return redirect(url_for("admin_home"))


@app.route("/admin/grant_premium", methods=["POST"])
@login_required
@admin_required
def admin_grant_premium():
    email = (request.form.get("email") or "").strip().lower()
    days = request.form.get("days", type=int)

    if not email or not days or days <= 0:
        return redirect(url_for("admin_home"))

    u = Usuario.query.filter_by(email=email).first()
    if not u:
        return redirect(url_for("admin_home"))

    base_dt = u.free_premium_until if u.free_premium_until and u.free_premium_until > now_utc() else now_utc()
    u.free_premium_until = base_dt + timedelta(days=days)
    db.session.commit()

    return redirect(url_for("admin_users"))


# =====================================================
# RUN
# =====================================================

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
